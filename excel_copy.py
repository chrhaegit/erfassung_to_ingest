import sys
import shutil
from openpyxl import load_workbook
from pathlib import Path

class ExcelMapping:
    def __init__(self, inputfile, destinationfile):
        self.wb_input = load_workbook(inputfile, data_only=True)
        self.ingest_path = destinationfile
        self.wb_ingest = load_workbook(destinationfile)

    def colindex_bycolname(self, currworksheet, colname):
        retcolindex = 0

        for colindex in range(1, 50):  # only looks up in the first 50 cols 
            header_name = currworksheet.cell(row=1, column=colindex).value
            if header_name == colname:
                retcolindex = colindex
                break
        return retcolindex
    
    def map(self, src_colname, src_rownr, dest_colname, dest_rownr) -> bool:
        src_colind = self.colindex_bycolname(self.wb_input["data sheet"], src_colname)
        if not src_colind:
            print(f"Error: No source header[{src_colname}] found in the input.xlsx")
            return False

        dest_colind = self.colindex_bycolname(self.wb_ingest["data sheet"], dest_colname)            
        if not dest_colind:
            print(f"Error: No destination header[{dest_colname}] found in the ingest-template")
            return False
        
        source_cell_value = self.wb_input["data sheet"].cell(src_rownr, src_colind).value
        self.wb_ingest["data sheet"].cell(dest_rownr, dest_colind).value = source_cell_value
        self.wb_ingest.save(self.ingest_path)
        return True                


    def do_mappings(self, test_mappings):        
        for src_colname, src_rownr, dest_colname, dest_rownr in test_mappings:               
            if not self.map(src_colname, src_rownr, dest_colname, dest_rownr):               
                return False
            
        return True

def test_mappings():
    # Define mappings of header names from inputfile.xlsx to ingest.xlsx
    return [
        ('Name', 2, 'Name', 4),
        ('Name', 2, 'Name2', 4),
        ('Vorname', 2, 'Vorname', 4)
    ]

def filehandling(input_path, template_path, ingest_path):
    if not input_path.exists():
        print(f"Error: File '{input_path}' not found.")
        return False
    if ingest_path.exists():
        ingest_path.unlink()
    
    if not template_path.exists():
        print("Error: templates.xlsx file not found.")
        return False
    
    shutil.copy(template_path, ingest_path)
    print(f"Copied {template_path} to {ingest_path}")
    return True

def main(inputfile):
    print("*" *20, "  START  ", "*" *20)
    if not filehandling(Path(inputfile), Path('./data/templates.xlsx'), Path('./ingest.xlsx')):
        return
    
    mapping = ExcelMapping(inputfile, Path('./ingest.xlsx'))
    if mapping.do_mappings(test_mappings()):
        print(f"sucessfully mapped to ingest-excel!")
    print("*" *20, "  ENDE  ", "*" *20)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <inputfile>")
        sys.exit(1)

    inputfile = sys.argv[1]
    main(inputfile)
