import sys
import shutil
from openpyxl import load_workbook
from pathlib import Path

class IngestExcelMapping:
    def __init__(self, inputfile, destinationfile):
        self.wb_input = load_workbook(inputfile, data_only=True)
        self.ingest_path = destinationfile
        self.wb_ingest = load_workbook(destinationfile)

    def colindex_bycolname(self, currworksheet, colname, header_row):
        retcolindex = 0

        for colindex in range(1, 50):  # only looks up in the first 50 cols 
            header_name = currworksheet.cell(row=header_row, column=colindex).value
            if header_name == colname:
                retcolindex = colindex
                break
        return retcolindex

    def map_vonbis_zeitraum(self, src_rownr, dest_rownr):
        src_von_colind = self.colindex_bycolname(self.wb_input["data sheet"], "von", 3 )
        if not src_von_colind:
            print(f"Error 10: No source header[von] found in the input.xlsx")
            return False    
        src_bis_colind = self.colindex_bycolname(self.wb_input["data sheet"], "bis", 3)
        if not src_bis_colind:
            print(f"Error 11: No source header[bis] found in the input.xlsx")
            return False     
 
        dest_colind = self.colindex_bycolname(self.wb_ingest["data sheet"], "Zeitraum", 1)            
        if not dest_colind:
            print(f"Error 12: No destination header[Zeitraum] found in the ingest-template")
            return False     

        src_von_value = str(self.wb_input["data sheet"].cell(src_rownr, src_von_colind).value) 
        src_bis_value = str(self.wb_input["data sheet"].cell(src_rownr, src_bis_colind).value)
        zeitr = src_von_value + " bis " + src_bis_value
        self.wb_ingest["data sheet"].cell(dest_rownr, dest_colind).value = zeitr
        self.wb_ingest.save(self.ingest_path)

    def map(self, src_colname, src_rownr, dest_colname, dest_rownr) -> bool:
        src_colind = self.colindex_bycolname(self.wb_input["data sheet"], src_colname, 3)
        if not src_colind:
            print(f"Error 01: No source header[{src_colname}] found in the input.xlsx")
            return False

        dest_colind = self.colindex_bycolname(self.wb_ingest["data sheet"], dest_colname, 1)            
        if not dest_colind:
            print(f"Error 02: No destination header[{dest_colname}] found in the ingest-template")
            return False
        
        source_cell_value = self.wb_input["data sheet"].cell(src_rownr, src_colind).value
        self.wb_ingest["data sheet"].cell(dest_rownr, dest_colind).value = source_cell_value
        self.wb_ingest.save(self.ingest_path)
        return True                


    def do_mappings(self, test_mappings):    
        for i in range(0, 6):    
            for src_colname, src_rownr, dest_colname, dest_rownr in test_mappings:               
                if not self.map(src_colname, src_rownr+i, dest_colname, dest_rownr+i):               
                    return False            
        return True

def col_mappings():
    # Define mappings of header names from inputfile.xlsx to ingest.xlsx
    return [
        ('Name', 4, 'Name', 2),
        ('Vorname', 4, 'Vorname', 2),
        ('Strasse', 4, 'Strasse', 2)
    ]

def filehandling(input_path, template_path, ingest_path):
    if not input_path.exists():
        print(f"Error: File '{input_path}' not found.")
        return False
    if ingest_path.exists():
        ingest_path.unlink()
    
    if not template_path.exists():
        print("Error: templates.xlsx file not found.")
        return False
    
    shutil.copy(template_path, ingest_path)
    print(f"Copied {template_path} to {ingest_path}")
    return True

def main(inputfile):
    print("*" *20, "  START  ", "*" *20)
    if not filehandling(Path(inputfile), Path('./data/templates.xlsx'), Path('./ingest.xlsx')):
        return
    
    mapping = IngestExcelMapping(inputfile, Path('./ingest.xlsx'))
    if mapping.do_mappings(col_mappings()):
        print(f"sucessfully created ingest-excel!")
    print("*" *20, "  ENDE  ", "*" *20)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <inputfile>")
        sys.exit(1)

    inputfile = sys.argv[1]
    main(inputfile)
