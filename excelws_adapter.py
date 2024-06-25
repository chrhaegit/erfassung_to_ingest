import openpyxl

class ExcelWSAdapter:

    def __init__(self, filename:str, worksheetname:str, read_only=False):
        self.filename = filename
        self.worksheetname = worksheetname
        self.wb = openpyxl.load_workbook(filename, read_only=read_only, data_only=True)
        self.ws = self.wb[worksheetname]

        self.col_header_row = self.ws.min_row
        self.firstdata_row = self.col_header_row + 1
        self.lastdata_row =  self.ws.max_row
        self.firstdata_col = self.ws.min_column
        self.lastdata_col = self.ws.max_column

        self.col_headers = {cell.value : cell.col_idx  for cell in self.ws[self.col_header_row]}        

    def mandatorycolumns(self):
        return {cell.value : cell.col_idx  for cell in self.ws[self.col_header_row] if cell.value[-1] == "*"}
    
    def checkmandatorycells(self)->bool:
        no_errors = True
        for row_index in range(self.firstdata_row, self.lastdata_row):
            for colname, col_index in self.mandatorycolumns().items(): 
                #print(row_index, col_index)
                if not self.ws.cell(row_index, col_index).value:
                    no_errors = False
                    print(f"[{colname}] ist ein Pflichtfeld, Zeile [{row_index}] ist leer!")
        return no_errors
    
    def checkcolvalues(self, colname:str, validatefunc):       
        tests_allok = True
        col_index = self.col_headers[colname]
        for row_index in range(self.firstdata_row, self.lastdata_row):
            val = self.ws.cell(row_index, col_index).value
            tests_allok = validatefunc(val, row_index)
        return tests_allok
    
    def cell(self, rownr:int, colname:str):
        colnr = self.col_headers[colname]
        return self.ws.cell(row=rownr, column=colnr)

    def setvalue(self, rownr:int, colname:str, newval):
        colnr = self.col_headers[colname]
        self.ws.cell(row=rownr, column=colnr).value = newval

    def save(self):
        self.wb.save(self.filename)

    def print_data(self):
        for row in self.ws.iter_rows(min_row=self.firstdata_row):
            for cell in row:
                print(cell.value)
            print("*" * 20)

    def __repr__(self) -> str:
        return f"{self.__class__.__name__}({self.filename}[{self.worksheetname}])"



def alter(x, row_index):
    if x is None:
        return False
    
    x = int(x)
    if x > 120:
        print(f"Alter [{x}] darf nicht grösser als 50 sein in Zeile [{row_index}]")
        return False
    return True

def process_cell(colname, erfassung_val):
    return erfassung_val

def copy_erfassungxl_to_ingestxl():
    print("*"*75)
    print("START vom script: Erfassung --> Ingest Excel:")
    print("*"*75)
    erfassxl = ExcelWSAdapter("Erfassung.xlsx", "data sheet")
    ingestxl = ExcelWSAdapter("ingest.xlsx", "data sheet")

    if not erfassxl.checkmandatorycells():
        print("Überprüfung der Pflichtfelder: Fehler --> Abbruch!")
        return
    else:
        print("Überprüfung der Pflichtfelder erfolgreich abgeschlossen")
        print("-"*75)

    alterfunc = alter
    if not erfassxl.checkcolvalues("Alter", alterfunc): 
        print("Überprüfung der Spalte Alte: Fehler --> Abbruch!")
        return
    else:
        print("Überprüfung der Spalte Alter erfolgreich abgeschlossen")
        print("-"*75)

    print("Start der reihenweisen Übertragung von Erfassung --> Ingest Excel: ")
    for row_index in range(erfassxl.firstdata_row, erfassxl.lastdata_row):             
        for colname, col_index in ingestxl.col_headers.items(): 
            cell_val = erfassxl.ws.cell(row_index, col_index).value
            pval = process_cell(colname, cell_val) #--> mit Exception werfen was machen
            #if not pval:
            #    print(f"Übertragung gestoppt bei (Zeile|Spalte): ({row_index}|{colname})")
            #    return
            #else:            
            ingestxl.cell(row_index, colname).value = pval
        print(f"Übertragung von Zeile {row_index} erfolgreich ....")

    ingestxl.save()
    print("*"*75)
    print("Übertragung vollständig abgeschlossen!")
    print("*"*75)

def main():
    xladapter = ExcelWSAdapter("Erfassung.xlsx", "data sheet")
    
    xladapter.checkmandatorycells()
    


    alterfunc = alter
    xladapter.checkcolvalues("Alter", alterfunc)

    #xladapter.setvalue(5, "Strasse", "Aarepark 2b")
    xladapter.cell(5, "Strasse").value = "Gugugesli 6"

    xladapter.save()

    #xladapter.print_data()
    
    # *******************************************
    # loop over colheader and set values in row x
    # *******************************************
    # x = 4
    # for colind, colname  in xladapter.col_headers.items():
    #     print(f"ind={colind:2}->name={colname}")
    #     value = "test"  # get the value from somewhere
    #     xladapter.setvalue(x, colname, value)
    

if __name__ == "__main__":
    #main()
    copy_erfassungxl_to_ingestxl()